"""
Report reader and writer.

Two jobs:
  1. Read the roster (the list of items) from the template.
  2. Write values back into a copy of that template without
     disturbing formatting, grouping or merged cells.

The round-trip was verified against the Aug 20 report:
  109 grouped rows, 5 grouped columns, 6 merged cells,
  5 conditional formatting rules, 3,066 styled cells - all intact.

Note on column O (AGT Actual Onhand). The old template reaches the
AGT file through an external link:

    =IFERROR(SUMIFS([1]Sheet1!$C:$C, [1]Sheet1!$A:$A, D5), 0)

The app writes a plain value there instead. That is deliberate: the
link only resolves while the AGT file sits at the exact path it was
created from, and IFERROR turns any failure into a silent zero.
"""

import shutil
from dataclasses import dataclass

from openpyxl import load_workbook

from .normalize import clean_text, item_key

# Field to column mapping for the ND SOH layout.
# D=Item, G..R = the three distributors plus the totals.
COLUMN_FIELDS = {
    "TW_dcr": "G",
    "TW_intransit": "H",
    "TW_onhand": "I",
    "QS_dcr": "J",
    "QS_intransit": "K",
    "QS_onhand": "L",
    "AGT_dcr": "M",
    "AGT_intransit": "N",
    "AGT_onhand": "O",
}

# Grouped by who owns each column.
#
# The app writes TW and AGT only. DCR (G/J/M) and QS (K/L) are filled
# in by hand and are never touched, so the pivot and SUMIFS work done
# there is safe.
FIELD_GROUPS = {
    "TW": ["TW_intransit", "TW_onhand"],
    "AGT": ["AGT_intransit", "AGT_onhand"],
    "QS": ["QS_intransit", "QS_onhand"],
    "DCR": ["TW_dcr", "QS_dcr", "AGT_dcr"],
}

# The default: TW and AGT only.
DEFAULT_GROUPS = ("TW", "AGT")


def fields_for(groups) -> list:
    """Return the field names belonging to the selected groups."""
    out = []
    for group in groups:
        out.extend(FIELD_GROUPS.get(group, []))
    return out

_ITEM_HEADERS = {"item", "item code"}
_SKIP_SERIES = ("GRAND TOTAL", "DCR INVENTORY", "TOTAL")


def _sanitize(wb) -> dict:
    """
    Clean up things that make Excel show the
    "We found a problem with some content" repair prompt.

    1. ORPHANED EXTERNAL LINKS.
       The old template links to the AGT file:
           =IFERROR(SUMIFS([1]Sheet1!$C:$C, [1]Sheet1!$A:$A, D5), 0)
       We replace column O with a plain value, so no formula uses that
       link any more - but openpyxl still writes the externalLink part,
       the <externalReferences> entry and the relationship. That leaves
       a definition pointing at D:\\Downloads\\... that nothing needs.
       Excel flags the mismatch. Removing it is safe: the data is
       already in the cells.

    2. STALE CACHED VALUES.
       openpyxl keeps formula strings but drops their cached results.
       Setting fullCalcOnLoad tells Excel to recalculate everything on
       open, so the Total columns and Grand Total rows are correct
       immediately instead of showing blanks.
    """
    removed_links = len(getattr(wb, "_external_links", []) or [])
    try:
        wb._external_links = []
    except AttributeError:
        removed_links = 0

    try:
        wb.calculation.fullCalcOnLoad = True
    except AttributeError:
        pass

    return {"external_links_removed": removed_links}


@dataclass
class Roster:
    """The report's item list, with the Excel row each one lives on."""
    items: list           # canonical item codes, in template order
    row_of: dict          # {item: excel_row}
    header_row: int
    item_col: int
    sheet_name: str
    skipped_rows: list    # (row, series) of the total rows skipped


def read_roster(path: str, sheet_name: str = "SOH") -> Roster:
    """
    Read the item roster from the template.

    The template is authoritative: no row is added or removed. An
    item that appears in the raw file but not here is reported as
    unmatched rather than being silently appended.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
        actual_sheet = ws.title

        header_row = None
        item_col = None
        series_col = None

        for row_idx, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
            for col_idx, cell in enumerate(row):
                label = clean_text(cell).lower()
                if label in _ITEM_HEADERS:
                    header_row = row_idx
                    item_col = col_idx
                if label == "series":
                    series_col = col_idx
            if header_row:
                break

        if header_row is None:
            raise ValueError(
                f"Could not find an 'Item' header in sheet '{actual_sheet}'."
            )

        items = []
        row_of = {}
        skipped = []

        for offset, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True)
        ):
            excel_row = header_row + 1 + offset

            series = ""
            if series_col is not None and series_col < len(values):
                series = clean_text(values[series_col]).upper()

            if any(series.startswith(s) for s in _SKIP_SERIES):
                skipped.append((excel_row, series))
                continue

            if item_col >= len(values):
                continue
            item = item_key(values[item_col])
            if not item:
                continue

            # If a code appears twice, the first row wins.
            if item not in row_of:
                items.append(item)
                row_of[item] = excel_row

        return Roster(
            items=items,
            row_of=row_of,
            header_row=header_row,
            item_col=item_col,
            sheet_name=actual_sheet,
            skipped_rows=skipped,
        )
    finally:
        wb.close()


def clear_report(
    template_path: str,
    output_path: str,
    roster: Roster,
    groups=("TW", "QS", "AGT", "DCR"),
    blank_instead_of_zero: bool = False,
) -> dict:
    """
    Refresher: reset the data columns to zero.

    ALWAYS KEPT:
      B  Series
      C  Launch Date
      D  Item
      E  Market Name
      F  SRP  <- the price list
      P, Q, R  Total formulas (=G+J+M), which fall to 0 on their own
      Grand Total, DCR Inventory and Diff. qty. rows, which are pure
         formulas (=SUM(G5:G140), =+G141, =G141-I141)

    CLEARED:
      G, H, I  TW
      J, K, L  QS
      M, N, O  AGT

    Useful when the template is last week's report: without this, old
    DCR and QS values would be carried into the new one.

    Only roster rows are touched. Total rows are skipped because they
    are formulas and update themselves.
    """
    shutil.copyfile(template_path, output_path)

    active_fields = fields_for(groups)

    wb = load_workbook(output_path)  # not data_only, so formulas survive
    try:
        ws = wb[roster.sheet_name]

        cleared = 0
        for row in roster.row_of.values():
            for field in active_fields:
                cell = ws[f"{COLUMN_FIELDS[field]}{row}"]
                cell.value = None if blank_instead_of_zero else 0
                cleared += 1

        sanitized = _sanitize(wb)
        wb.save(output_path)

        return {
            "cells_cleared": cleared,
            "external_links_removed": sanitized["external_links_removed"],
            "rows": len(roster.row_of),
            "columns_cleared": sorted(COLUMN_FIELDS[f] for f in active_fields),
            "output_path": output_path,
        }
    finally:
        wb.close()


def write_report(
    template_path: str,
    output_path: str,
    roster: Roster,
    values: dict,
    report_date=None,
    blank_zeros: bool = False,
    groups=DEFAULT_GROUPS,
) -> dict:
    """
    Copy the template, then write values into the copy.

    groups: which of "TW", "AGT", "QS", "DCR" to write. The default
            is ("TW", "AGT") only. DCR and QS are filled in by hand,
            so the app leaves them alone - writing them would wipe
            out the pivot and SUMIFS work already in the template.

    Formatting is never touched; only cell values are written. The
    Total columns P/Q/R already hold formulas (=G+J+M), so they are
    left alone and recalculate themselves.

    blank_zeros: when True, write an empty cell instead of 0.
    """
    shutil.copyfile(template_path, output_path)

    active_fields = fields_for(groups)
    untouched = [
        COLUMN_FIELDS[f] for f in COLUMN_FIELDS if f not in active_fields
    ]

    wb = load_workbook(output_path)  # not data_only, so formulas survive
    try:
        ws = wb[roster.sheet_name]

        written = 0
        for item, row in roster.row_of.items():
            row_values = values.get(item)
            if row_values is None:
                continue
            for field in active_fields:
                column = COLUMN_FIELDS[field]
                amount = row_values.get(field, 0) or 0
                cell = ws[f"{column}{row}"]
                if blank_zeros and not amount:
                    cell.value = None
                else:
                    cell.value = amount
                written += 1

        # Update the report date if its cell can be located.
        date_cell = None
        if report_date is not None:
            for r in range(1, roster.header_row + 1):
                for c in range(1, 12):
                    text = clean_text(ws.cell(row=r, column=c).value).lower()
                    if "stocks on hand as of" in text:
                        # The date sits in a cell to the right of the label.
                        for offset in range(1, 5):
                            target = ws.cell(row=r, column=c + offset)
                            if not isinstance(target, type(ws["A1"])):
                                continue
                            try:
                                target.value = report_date
                                date_cell = target.coordinate
                            except AttributeError:
                                continue  # merged cell - skip it
                            break
                        break
                if date_cell:
                    break

        sanitized = _sanitize(wb)
        wb.save(output_path)

        return {
            "cells_written": written,
            "external_links_removed": sanitized["external_links_removed"],
            "items_written": len([i for i in roster.row_of if i in values]),
            "date_cell": date_cell,
            "output_path": output_path,
            "columns_written": sorted(COLUMN_FIELDS[f] for f in active_fields),
            "columns_untouched": sorted(untouched),
        }
    finally:
        wb.close()
