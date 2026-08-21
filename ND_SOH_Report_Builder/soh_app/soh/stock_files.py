"""
Parser para sa TW at AGT raw stock files.

Pareho ang layout ng dalawang file na ito (na-verify laban sa
TW Aug 20 at AGT Aug 20):

    row 3 (Excel) = header:  Series | Item | Market Name | Model | Intransit | Onhand
    row 4 pataas   = data
    may "Grand Total:" row sa dulo na dapat i-exclude

Hindi naka-hardcode ang column positions - hinahanap natin ang header
row sa pamamagitan ng pangalan, para hindi masira kapag nagdagdag ng
column ang system nila.
"""

from collections import defaultdict
from dataclasses import dataclass, field

from openpyxl import load_workbook

from .normalize import clean_text, item_key, is_ldu, describe_issues

# Mga posibleng pangalan ng bawat column, lowercase.
_HEADER_ALIASES = {
    "series": {"series"},
    "item": {"item", "item code", "itemcode"},
    "market": {"market name", "market", "marketname"},
    "model": {"model"},
    "intransit": {"intransit", "in transit", "in-transit"},
    "onhand": {"onhand", "on hand", "on-hand", "actual onhand"},
}


@dataclass
class StockRow:
    """Isang row mula sa raw file, bago pa i-merge."""
    excel_row: int
    series: str
    raw_item: str
    item: str          # canonical key (nalinis, tanggal ang LDU)
    market: str
    intransit: float
    onhand: float
    was_ldu: bool
    issues: list = field(default_factory=list)


@dataclass
class StockFile:
    """Resulta ng pag-parse ng isang raw stock file."""
    path: str
    sheet_name: str
    rows: list                     # list[StockRow] - bago mag-merge
    totals_row: tuple = None       # (intransit, onhand) mula sa Grand Total row
    header_row: int = None

    def merged(self) -> dict:
        """
        I-merge ang lahat ng row per canonical item code.
        Dito nangyayari ang LDU merge at ang pagsasama ng duplicate rows.

        Returns: {item_code: {"intransit": x, "onhand": y}}
        """
        out = defaultdict(lambda: {"intransit": 0.0, "onhand": 0.0})
        for row in self.rows:
            out[row.item]["intransit"] += row.intransit
            out[row.item]["onhand"] += row.onhand
        return dict(out)

    def item_totals(self) -> tuple:
        """Kabuuang Intransit at Onhand mula sa item rows (walang Grand Total)."""
        return (
            sum(r.intransit for r in self.rows),
            sum(r.onhand for r in self.rows),
        )

    def totals_match(self) -> bool:
        """
        Ito ang Step 6 sa workflow doc: 'TW raw total = TW Pivot total'.

        Kung may Grand Total row ang file, ikinukumpara natin ang item-sum
        dito. Kapag hindi tugma, may nawawalang row o may nabilang nang doble.
        """
        if self.totals_row is None:
            return True
        item_i, item_o = self.item_totals()
        gt_i, gt_o = self.totals_row
        return abs(item_i - gt_i) < 0.5 and abs(item_o - gt_o) < 0.5

    def rows_with_issues(self) -> list:
        return [r for r in self.rows if r.issues]


def _to_number(value) -> float:
    """
    Gawing number ang isang cell value.

    Hinahandle ang mga kaso sa Step 4 ng workflow doc: blank, dash,
    at 'number stored as text'.
    """
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)

    text = clean_text(value)
    if text in ("", "-", "--", "N/A", "n/a"):
        return 0.0

    # Tanggalin ang thousands separator at currency-style parentheses.
    text = text.replace(",", "").replace("(", "-").replace(")", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def _find_header(ws, max_scan: int = 15):
    """
    Hanapin ang header row at ang column index ng bawat field.

    Ibinabalik: (excel_row_number, {field: zero_based_col_index})
    """
    for row_idx, row in enumerate(ws.iter_rows(max_row=max_scan, values_only=True), start=1):
        labels = {}
        for col_idx, cell in enumerate(row):
            label = clean_text(cell).lower()
            if not label:
                continue
            for field_name, accepted in _HEADER_ALIASES.items():
                if label in accepted and field_name not in labels:
                    labels[field_name] = col_idx

        # Kailangan man lang may Item at isa sa dalawang quantity columns.
        if "item" in labels and ("intransit" in labels or "onhand" in labels):
            return row_idx, labels

    raise ValueError(
        "Could not find the header row. Expected columns named "
        "'Item', 'Intransit' and 'Onhand' within the first 15 rows."
    )


def _pick_sheet(wb) -> str:
    """
    Piliin ang sheet na may totoong SOH data.

    Ang AGT file ay minsan may dalawang sheet: ang raw na
    'SOH as of MM.DD.YYYY' at ang pivot na 'Sheet1' (Row Labels /
    Sum of Intransit / Sum of Onhand). Ang pivot ang unang sheet,
    kaya hindi puwedeng basta kunin ang wb.sheetnames[0].

    Ang hinahanap natin ay ang sheet na may tunay na header:
    Item, Intransit, Onhand. Hindi tumutugma ang pivot dahil
    'Row Labels' ang column header niya, hindi 'Item'.
    """
    candidates = []
    for name in wb.sheetnames:
        try:
            _find_header(wb[name])
        except ValueError:
            continue
        # Mas mataas ang priyoridad ng sheet na may "SOH" sa pangalan.
        score = 1 if "soh" in name.lower() else 0
        candidates.append((score, name))

    if not candidates:
        raise ValueError(
            "No sheet in this file has 'Item', 'Intransit' and 'Onhand' "
            "headers. If the file only contains a pivot, the raw SOH sheet "
            "is required."
        )

    candidates.sort(key=lambda x: -x[0])
    return candidates[0][1]


def parse_stock_file(path: str, sheet_name: str = None) -> StockFile:
    """
    Basahin ang isang TW o AGT raw file.

    Kung walang ibinigay na sheet_name, automatic na hahanapin ang
    tamang sheet - hindi basta ang una, dahil minsan pivot ang una.
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb[_pick_sheet(wb)]
        actual_sheet = ws.title

        header_row, cols = _find_header(ws)

        rows = []
        totals_row = None

        for offset, values in enumerate(
            ws.iter_rows(min_row=header_row + 1, values_only=True)
        ):
            excel_row = header_row + 1 + offset

            def get(field_name):
                idx = cols.get(field_name)
                if idx is None or idx >= len(values):
                    return None
                return values[idx]

            series = clean_text(get("series"))
            raw_item = get("item")
            item = item_key(raw_item)

            # Hulihin ang Grand Total row - huwag isama sa data,
            # pero itago para magamit sa validation.
            joined = f"{series} {clean_text(raw_item)}".upper()
            if "GRAND TOTAL" in joined or joined.strip().startswith("TOTAL"):
                totals_row = (
                    _to_number(get("intransit")),
                    _to_number(get("onhand")),
                )
                continue

            if not item:
                continue

            intransit = _to_number(get("intransit"))
            onhand = _to_number(get("onhand"))

            rows.append(
                StockRow(
                    excel_row=excel_row,
                    series=series,
                    raw_item=clean_text(raw_item),
                    item=item,
                    market=clean_text(get("market")),
                    intransit=intransit,
                    onhand=onhand,
                    was_ldu=is_ldu(raw_item),
                    issues=describe_issues(raw_item),
                )
            )

        return StockFile(
            path=path,
            sheet_name=actual_sheet,
            rows=rows,
            totals_row=totals_row,
            header_row=header_row,
        )
    finally:
        wb.close()
